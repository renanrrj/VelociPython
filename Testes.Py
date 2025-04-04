from selenium import webdriver # primeira linha para importar o webdriver
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from datetime import datetime
import time    
from openpyxl import Workbook
import openpyxl as opx
 
servico = Service(GeckoDriverManager().install()) 
navegador = webdriver.Firefox(service=servico)

navegador.get('https://www.speedtest.net/pt')
time.sleep(3)

navegador.find_element (By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[1]/a/span[4]').click()
time.sleep(40)

Data = datetime.today().strftime('%Y-%m-%d')
Hora = datetime.now().time().strftime('%H:%M:%S')
Download = navegador.find_element (By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[1]/div[1]/div/div[2]/span').text
Upload = navegador.find_element (By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[1]/div[2]/div/div[2]/span').text
Min = int(navegador.find_element (By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[2]/div/span[3]/span' ).text)
Max =  int (navegador.find_element (By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[2]/div/span[4]/span' ).text)
Media = int(navegador.find_element (By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[2]/div/span[2]/span' ).text)

plan = ""

Meses = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
NMes = datetime.now().month
MesCorrente =  Meses[NMes]
MesPassado = Meses[NMes - 1]

try:               
    plan = opx.load_workbook(filename = 'Z:/7_Geral (RENAN)/Informática/Monitoramento de Link.xlsx')
    if MesCorrente != MesPassado:        
        plan.create_sheet(MesCorrente)
    pag_med = plan[MesCorrente] # Planilha PLAN e aba MESCORRENTE
    pag_med.append([Data, Hora, "", Download, Upload, "", Min, Max, Media])

except:    
    plan = opx.Workbook()
    plan.active.title = MesCorrente
    pag_med = plan.active
    pag_med.append(['Data', 'Hora', '', 'Download', 'Upload','','Min', 'Max', 'Med'])
    pag_med.append([Data, Hora, "", Download, Upload, "", Min, Max, Media])

plan.save('Z:/7_Geral (RENAN)/Informática/Monitoramento de Link.xlsx')
navegador.close()