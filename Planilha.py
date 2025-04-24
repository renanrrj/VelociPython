import openpyxl as opx
from datetime import datetime

def obter_mes_atual():
    Meses = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    NMesAtual = datetime.now().month
    return Meses[NMesAtual]

def carregar_ou_criar_planilha(caminho, MesPlanilha):
    try:
        plan = opx.load_workbook(filename=caminho)
        if MesPlanilha not in plan.sheetnames:
            pag_med = plan.create_sheet(MesPlanilha)
        else:
            pag_med = plan[MesPlanilha]
    except FileNotFoundError:
        plan = opx.Workbook()
        plan.active.title = MesPlanilha
        pag_med = plan.active
        pag_med.append(['Data', 'Hora', 'Download', 'Upload', 'Min', 'Max', 'Med'])

    return plan, pag_med

def salvar_dados(plan, pag_med, Data, Hora, Download, Upload, Min, Max, Media, caminho):
    pag_med.append([Data, Hora, Download, Upload, Min, Max, Media])
    plan.save(caminho)
