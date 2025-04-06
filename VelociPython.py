import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.firefox import GeckoDriverManager
from datetime import datetime
import openpyxl as opx
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Inicia o serviço do GeckoDriver e o navegador
servico = Service(GeckoDriverManager().install())
navegador = webdriver.Firefox(service=servico)

# Acessa o site Speedtest
navegador.get('https://www.speedtest.net/pt')

# Espera até que o botão de iniciar o teste esteja visível
WebDriverWait(navegador, 10).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[1]/a/span[4]')))

# Clica no botão para iniciar o teste
navegador.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[1]/a/span[4]').click()

# Espera o tempo necessário para o teste
time.sleep(40)

# Obtém os dados do teste
Data = datetime.today().strftime('%Y-%m-%d')
Hora = datetime.now().time().strftime('%H:%M:%S')

# Pegando os valores de download, upload, min, max e media
Download = navegador.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[1]/div[1]/div/div[2]/span').text
Upload = navegador.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[1]/div[2]/div/div[2]/span').text
Min = int(navegador.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[2]/div/span[3]/span').text)
Max = int(navegador.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[2]/div/span[4]/span').text)
Media = int(navegador.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div[3]/div/div/div[2]/div[2]/div/span[2]/span').text)

# Mapeamento dos meses em português
Meses = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
NMesAtual = datetime.now().month
MesPlanilha = Meses[NMesAtual]

# Tenta carregar a planilha existente ou cria uma nova
try:
    plan = opx.load_workbook(filename='C:/Users/rNz/Desktop/planilha.xlsx')
    if MesPlanilha in plan.sheetnames:
        pag_med = plan[MesPlanilha]
    else:
        pag_med = plan.create_sheet(MesPlanilha)
    
    # Adiciona os dados na planilha
    pag_med.append([Data, Hora, "", Download, Upload, "", Min, Max, Media])

except FileNotFoundError:
    # Se o arquivo não for encontrado, cria um novo
    plan = opx.Workbook()
    plan.active.title = MesPlanilha
    pag_med = plan.active

    # Adiciona o cabeçalho
    pag_med.append(['Data', 'Hora', '', 'Download', 'Upload', '', 'Min', 'Max', 'Med'])

    # Adiciona a linha de dados
    pag_med.append([Data, Hora, "", Download, Upload, "", Min, Max, Media])

# Salva a planilha
plan.save('C:/Users/rNz/Desktop/planilha.xlsx')

# Fecha o navegador
navegador.quit()
